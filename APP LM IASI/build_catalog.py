import openpyxl
import json
import os

def build_catalog():
    print("Loading workbook...")
    wb = openpyxl.load_workbook('catalog.xlsx', read_only=True)
    sheet = wb.active
    
    catalog = {}
    print("Processing rows...")
    
    # iter_rows yields tuples
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows) # Skip header
    
    count = 0
    for row in rows:
        if not row[0]:
            continue # empty row
        
        cod = str(row[0]).strip()
        denumire = str(row[1]).strip() if row[1] else ""
        raion = str(row[2]).strip() if row[2] else ""
        ean = str(row[4]).strip() if row[4] else ""
        
        # Add entry for SKU
        if cod:
            catalog[cod] = {"d": denumire, "r": raion}
        
        # Add entry for EAN
        if ean and ean != "None" and ean != cod:
            catalog[ean] = {"d": denumire, "r": raion, "sku": cod}
            
        count += 1
        if count % 10000 == 0:
            print(f"Processed {count} rows...")
            
    print(f"Total processed rows: {count}")
    print(f"Total dictionary keys: {len(catalog)}")
    
    with open('catalog.json', 'w', encoding='utf-8') as f:
        json.dump(catalog, f, separators=(',', ':'))
        
    print(f"Done! Created catalog.json ({os.path.getsize('catalog.json') / 1024 / 1024:.2f} MB)")

if __name__ == "__main__":
    build_catalog()
